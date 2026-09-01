import json
import os
import shutil
import threading
import traceback
import webbrowser
from datetime import date, datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(BASE_DIR)
WORKBOOK_PATH = os.path.join(ROOT_DIR, "Socios_2026-08-31.xlsx")
ARCHIVE_DIR = os.path.join(ROOT_DIR, "Arquivo")
YEARS = [2026, 2027, 2028, 2029, 2030]
FIELDS = ["name", "contact", "nif", "locality", "address", "postal", "email", "paymentMode", "date", "notes"]
REMOVED_MARKER = "[REMOVIDO]"


def excel_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return "" if value is None else str(value)


def read_members():
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    sheet = workbook["Listagem de Sócios"]
    members = []
    for row in range(4, 54):
        name = sheet.cell(row, 2).value
        if not name:
            continue
        notes = excel_value(sheet.cell(row, 17).value)
        removed = notes.startswith(REMOVED_MARKER)
        member = {
            "id": f"excel-{sheet.cell(row, 1).value}",
            "number": int(sheet.cell(row, 1).value),
            "name": excel_value(name),
            "contact": excel_value(sheet.cell(row, 3).value),
            "nif": excel_value(sheet.cell(row, 4).value),
            "locality": excel_value(sheet.cell(row, 5).value),
            "address": excel_value(sheet.cell(row, 6).value),
            "postal": excel_value(sheet.cell(row, 7).value),
            "email": excel_value(sheet.cell(row, 8).value),
            "paymentMode": excel_value(sheet.cell(row, 15).value),
            "date": excel_value(sheet.cell(row, 16).value),
            "notes": notes[len(REMOVED_MARKER):].lstrip(" -") if removed else notes,
            "removed": removed,
            "dues": {year: excel_value(sheet.cell(row, 10 + index).value) or "Pendente" for index, year in enumerate(YEARS)},
        }
        members.append(member)
    return members


def write_members(members):
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    archive_path = os.path.join(ARCHIVE_DIR, f"registo socios_{stamp}.xlsx")
    suffix = 2
    while os.path.exists(archive_path):
        archive_path = os.path.join(ARCHIVE_DIR, f"registo socios_{stamp}_{suffix}.xlsx")
        suffix += 1
    shutil.copy2(WORKBOOK_PATH, archive_path)
    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook["Listagem de Sócios"]
    for row in range(4, 54):
        for column in range(2, 18):
            sheet.cell(row, column).value = None
    for index, member in enumerate(members[:50], start=4):
        sheet.cell(index, 1).value = member.get("number", index - 3)
        values = [member.get(field, "") for field in FIELDS[:7]]
        for column, value in enumerate(values, start=2):
            sheet.cell(index, column).value = value
        notes = member.get("notes", "")
        if member.get("removed"):
            notes = f"{REMOVED_MARKER} - {notes}" if notes else REMOVED_MARKER
        for column, value in ((15, member.get("paymentMode", "")), (16, member.get("date", "")), (17, notes)):
            sheet.cell(index, column).value = value
        for due_index, year in enumerate(YEARS, start=10):
            sheet.cell(index, due_index).value = member.get("dues", {}).get(str(year), member.get("dues", {}).get(year, "Pendente"))
    workbook.save(WORKBOOK_PATH)


class AppHandler(SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/api/members":
            self.send_json(read_members())
            return
        super().do_GET()

    def do_POST(self):
        if self.path != "/api/members":
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", 0))
            members = json.loads(self.rfile.read(length).decode("utf-8"))
            write_members(members)
            self.send_json({"ok": True})
        except Exception as error:
            traceback.print_exc()
            self.send_json({"error": str(error)}, status=500)

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    os.chdir(BASE_DIR)
    server = ThreadingHTTPServer(("127.0.0.1", 8765), AppHandler)
    print("TEAM JM aberta em http://127.0.0.1:8765")
    threading.Timer(0.8, lambda: webbrowser.open("http://127.0.0.1:8765")).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        server.server_close()
